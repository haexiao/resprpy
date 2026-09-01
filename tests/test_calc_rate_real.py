"""calc_rate on the user's real OXY-10 data, compared against R outputs.

R reference (reference/gen_refs_calc_rate.R): reads the same xlsx files with
readxl, builds time = 'Delta T [min]' * 60, oxygen = 'Oxygen', and runs
respR::calc_rate with the same 20 segments as the user's saved rmr1.csv.
"""
import os

import numpy as np
import openpyxl

from resprpy.calc import calc_rate
from conftest import REF_DIR, assert_close

DATADIR = r"X:\Rtools\20260422\20260422 20"
FILES = [1, 2, 6]


def read_channel(xlsx_path):
    """Extract (time_s, oxygen) from the OXY-10 export's channel sheet."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        sheet = next(s for s in wb.sheetnames if "Ch " in s)
        ws = wb[sheet]
        rows = ws.iter_rows(values_only=True)
        header = next(rows)
        i_t = header.index("Delta T [min]")
        i_o = header.index("Oxygen")
        t, o = [], []
        for r in rows:
            if r[i_t] is None or r[i_o] is None:
                continue
            t.append(float(r[i_t]) * 60.0)
            o.append(float(r[i_o]))
    finally:
        wb.close()
    return np.column_stack([np.array(t), np.array(o)])


def _segments():
    seg = np.genfromtxt(os.path.join(DATADIR, "rmr1.csv"), delimiter=",",
                        skip_header=1, usecols=(6, 7))  # time, endtime
    return seg[:, 0], seg[:, 1]


def test_calc_rate_real_data_against_R():
    from_, to = _segments()
    for f in FILES:
        df = read_channel(os.path.join(DATADIR, f"{f}.xlsx"))
        cr = calc_rate(df, from_=from_, to=to, plot=False)
        ref = np.genfromtxt(os.path.join(REF_DIR, f"ref_calc_rate_file{f}.csv"),
                            delimiter=",", skip_header=1)

        s = cr.summary
        # column order in ref CSV: rank, intercept_b0, slope_b1, rsq, row,
        # endrow, time, endtime, oxy, endoxy, rate.2pt, rate
        assert_close(s["intercept_b0"], ref[:, 1], rtol=1e-8, label=f"file{f} intercept")
        assert_close(s["slope_b1"], ref[:, 2], rtol=1e-8, label=f"file{f} slope")
        assert_close(s["rsq"], ref[:, 3], rtol=1e-12, label=f"file{f} rsq")
        assert_close(s["row"], ref[:, 4], rtol=1e-12, label=f"file{f} row")
        assert_close(s["endrow"], ref[:, 5], rtol=1e-12, label=f"file{f} endrow")
        assert_close(s["time"], ref[:, 6], rtol=1e-12, label=f"file{f} time")
        assert_close(s["endtime"], ref[:, 7], rtol=1e-12, label=f"file{f} endtime")
        assert_close(s["oxy"], ref[:, 8], rtol=1e-9, label=f"file{f} oxy")
        assert_close(s["endoxy"], ref[:, 9], rtol=1e-9, label=f"file{f} endoxy")
        assert_close(s["rate.2pt"], ref[:, 10], rtol=1e-8, label=f"file{f} rate.2pt")
        assert_close(s["rate"], ref[:, 11], rtol=1e-8, label=f"file{f} rate")
        assert np.all(s["rank"] == ref[:, 0])


def test_calc_rate_matches_saved_rmr1():
    """The user's saved rmr1.csv (row 1) should be closely reproduced."""
    from_, to = _segments()
    df = read_channel(os.path.join(DATADIR, "1.xlsx"))
    cr = calc_rate(df, from_=from_, to=to, plot=False)
    s = cr.summary
    # R on the same xlsx (recipe B) gives slope -4.316843e-05, rsq 0.903
    assert np.isclose(s["slope_b1"][0], -4.316843e-05, rtol=1e-6)
    assert s["rsq"][0] == 0.903
    assert s["row"][0] == 157 and s["endrow"][0] == 432
